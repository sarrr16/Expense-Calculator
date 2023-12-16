from tkinter import *
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference, BarChart

root = Tk()
root.geometry("550x250")
root.title("Expense Tracker")

file_entry_label = Label(root, text="Enter file name:", font="Arial")
file_entry_label.place(x=250, y=0)
file_entry = Entry(root)
file_entry.place(x=250, y = 45)

def clear():
    textbox.delete("0","end")
    textbox2.delete("0","end")
    textbox.insert(0, str(0))
    textbox2.insert(0, str(0))

def upload():
            
        file_name = file_entry.get()
        file_path = f"C:/expenses/{file_name}.xlsx"

        try:
            wb = load_workbook(file_path)
        except FileNotFoundError:
            print("File not found. Creating a new file.")
            wb = Workbook()

        ws = wb.active

        data = ['Food', 'Travel', 'Data', 'Recharge', 'Study', 'Savings', 'Misc', 'Total']
        data1 = ['Rent', 'Maid', 'Electricity', 'Actual Total']
        data2 = ['Upi', 'Cash', 'Total']

        for i, value3 in enumerate(data2, 1):
            ws.cell(row=i+1, column=1).value = value3
            
        for i, value in enumerate(data, 1):
            ws.cell(row=1, column=i+1).value = value

        for i, value1 in enumerate(data1, 1):
            ws.cell(row=1, column=i+10).value = value1

        data = [float(textbox.get())]
        data1 = [float(textbox2.get())]

        for col in range(2, 9):
            sum_formula = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}3)"
            ws.cell(row=ws.max_row, column=col, value=sum_formula)

        for row in range(2, 5):
            sum_formula = f"SUM(B{row}:H{row})"
            ws.cell(row=row, column=9, value=f"={sum_formula}")

        for col in range(11, 14):
            sum_formula = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}3)"
            ws.cell(row=ws.max_row, column=col, value=sum_formula)

        for row in range(2, 5):
            sum_formula = f"SUM(I{row}:M{row})"
            ws.cell(row=row, column=14, value=f"={sum_formula}")

        def data333():
            if clicked.get() == "Food" :
                for i, value4 in enumerate(data, start=1):
                    existing_value = ws.cell(row=2, column=2).value or 0
                    ws.cell(row=2, column=2).value = existing_value + value4
                for i, value5 in enumerate(data1, start=1):
                    existing_value = ws.cell(row=3, column=2).value or 0
                    ws.cell(row=3, column=2).value = existing_value + value5
            elif clicked.get() == "Travel" :
                for i, value4 in enumerate(data, start=1):
                    existing_value = ws.cell(row=2, column=3).value or 0
                    ws.cell(row=2, column=3).value = existing_value + value4
                for i, value5 in enumerate(data1, start=1):
                    existing_value = ws.cell(row=3, column=3).value or 0
                    ws.cell(row=3, column=3).value = existing_value + value5
            elif clicked.get() == "Data" :
                for i, value4 in enumerate(data, start=1):
                    existing_value = ws.cell(row=2, column=4).value or 0
                    ws.cell(row=2, column=4).value = existing_value + value4
                for i, value5 in enumerate(data1, start=1):
                    existing_value = ws.cell(row=3, column=4).value or 0
                    ws.cell(row=3, column=4).value = existing_value + value5
            elif clicked.get() == "Recharge" :
                for i, value4 in enumerate(data, start=1):
                    existing_value = ws.cell(row=2, column=5).value or 0
                    ws.cell(row=2, column=5).value = existing_value + value4
                for i, value5 in enumerate(data1, start=1):
                    existing_value = ws.cell(row=3, column=5).value or 0
                    ws.cell(row=3, column=5).value = existing_value + value5
            elif clicked.get() == "Study" :
                for i, value4 in enumerate(data, start=1):
                    existing_value = ws.cell(row=2, column=6).value or 0
                    ws.cell(row=2, column=6).value = existing_value + value4
                for i, value5 in enumerate(data1, start=1):
                    existing_value = ws.cell(row=3, column=6).value or 0
                    ws.cell(row=3, column=6).value = existing_value + value5
            elif clicked.get() == "Savings" :
                for i, value4 in enumerate(data, start=1):
                    existing_value = ws.cell(row=2, column=7).value or 0
                    ws.cell(row=2, column=7).value = existing_value + value4
                for i, value5 in enumerate(data1, start=1):
                    existing_value = ws.cell(row=3, column=7).value or 0
                    ws.cell(row=3, column=7).value = existing_value + value5     
            elif clicked.get() == "Misc" :
                for i, value4 in enumerate(data, start=1):
                    existing_value = ws.cell(row=2, column=8).value or 0
                    ws.cell(row=2, column=8).value = existing_value + value4
                for i, value5 in enumerate(data1, start=1):
                    existing_value = ws.cell(row=3, column=8).value or 0
                    ws.cell(row=3, column=8).value = existing_value + value5
            elif clicked.get() == "Rent" :
                for i, value4 in enumerate(data, start=1):
                    existing_value = ws.cell(row=2, column=11).value or 0
                    ws.cell(row=2, column=11).value = existing_value + value4
                for i, value5 in enumerate(data1, start=1):
                    existing_value = ws.cell(row=3, column=11).value or 0
                    ws.cell(row=3, column=11).value = existing_value + value5
            elif clicked.get() == "Maid" :
                for i, value4 in enumerate(data, start=1):
                    existing_value = ws.cell(row=2, column=12).value or 0
                    ws.cell(row=2, column=12).value = existing_value + value4
                for i, value5 in enumerate(data1, start=1):
                    existing_value = ws.cell(row=3, column=12).value or 0
                    ws.cell(row=3, column=12).value = existing_value + value5
            elif clicked.get() == "Electricity" :
                for i, value4 in enumerate(data, start=1):
                    existing_value = ws.cell(row=2, column=13).value or 0
                    ws.cell(row=2, column=13).value = existing_value + value4
                for i, value5 in enumerate(data1, start=1):
                    existing_value = ws.cell(row=3, column=13).value or 0
                    ws.cell(row=3, column=13).value = existing_value + value5
            else:
                return 0
            
            pie = PieChart()
            labels = Reference(ws, min_col=2, min_row=1, max_col=8)
            data6 = Reference(ws, min_col=2, min_row=4, max_col=8)
            pie.add_data(data6, from_rows=True)
            pie.set_categories(labels)
            pie.title = "Expenses"
            ws.add_chart(pie, "A7")

            bar = BarChart()
            categories = Reference(ws, min_col=2, max_col=9, min_row=1)
            values = Reference(ws, min_col=2, max_col=9, min_row=2, max_row=3)
            bar.add_data(values, from_rows= True)
            bar.set_categories(categories)
            bar.title = "Expenses"
            ws.add_chart(bar, "J7")
            
        data333()
        clear()
        wb.save(file_path)

label1 = Label(root, text="UPI", font="Arial")
label1.place(x=5, y=45)
label2 = Label(root, text="Cash", font="Arial")
label2.place(x=5, y=90)
foodvar = DoubleVar()
food_var = DoubleVar() 
textbox = Entry(root, width=9,textvariable=foodvar)
textbox.place(x=60, y=45)
textbox2 = Entry(root, width=9,textvariable=food_var)
textbox2.place(x=60, y=90)
def update_label3(event, label3, clicked):
    label3.config(text=clicked.get())
    label3.place(x=60, y=5)

def go():
    if clicked.get() in ["Food", "Travel", "Data", "Recharge", "Savings", "Study", "Misc", "Electricity", "Rent", "Maid"]:
        upload()

    else:
        return 0

options = ['Food', 'Travel', 'Data', 'Recharge', 'Study', 'Savings', 'Misc', 'Rent', 'Maid', 'Electricity']

clicked = StringVar()
clicked.set("Food")
label3 = Label(root, text=clicked.get(), font="Arial")
label3.place(x=60, y=5)
drop = OptionMenu(root, clicked, *options,  command=lambda event: update_label3(event, label3, clicked))
drop.place(x=157, y=0)
button2 = Button(root, text="Upload", width=8, command=upload)
button2.place(x=157, y=45)


root.mainloop()